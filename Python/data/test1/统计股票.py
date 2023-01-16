from _datetime import datetime
# print(datetime.now().strftime('%Y年%m月%d日'))
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

# 加载一个工作pu--》Workbook
wb = openpyxl.load_workbook('resources/阿里巴巴2020年股票数据.xlsx')

sheet = wb.worksheets[0]
max=sheet.max_row
npx =np.array([x[0].value for x in sheet[2:max]])
print(type(npx[0]))
npy=np.array([x[6].value for x in sheet[2:max]])
# print(wbx)


plt.rcParams['figure.dpi'] = 300 #分辨率

x=np.linspace(sum(npy)/len(npy),sum(npy)/len(npy),num=max-1)
plt.plot(npx,npy,npx,x)
plt.show()
# print(sheet.dimensions)
# print(sheet.max_row, sheet.max_column)
