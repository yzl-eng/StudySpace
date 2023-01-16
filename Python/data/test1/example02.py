# from _datetime import datetime
# print(datetime.now().strftime('%Y年%m月%d日'))

import openpyxl

# 加载一个工作pu--》Workbook
wb = openpyxl.load_workbook('resources/阿里巴巴2020年股票数据.xlsx')
print(type(wb))

# 获取工作表名字
# sheet = wb['股票数据'] #无代码提示
sheet = wb.worksheets[0]
print(type(sheet))
print(sheet.dimensions)
print(sheet.max_row, sheet.max_column)
